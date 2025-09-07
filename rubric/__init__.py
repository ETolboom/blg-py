import io

import pandas as pd

from typing import Union, List, Optional

from pydantic import BaseModel

from algorithms import AlgorithmResult
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class Assignment(BaseModel):
    # Set default values in case we want to onboard without
    # a reference model/description
    reference_xml: Optional[str] = ""
    description: Optional[str] = ""


class RubricCriterion(AlgorithmResult):
    custom_score: Union[float, None]
    default_points: float


class OnboardingRubric(BaseModel):
    assignment: Assignment
    algorithms: List[str] = []


class Rubric(BaseModel):
    criteria: List[RubricCriterion]
    assignment: Union[Assignment, None]

    def to_excel_worksheet(self, writer, filename: str) -> None:
        workbook = writer.book
        worksheet = workbook.create_sheet(filename)

        # Define styles
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        criterion_font = Font(bold=True, size=11)
        criterion_fill = PatternFill(start_color='E8F1FF', end_color='E8F1FF', fill_type='solid')
        border = Border(
            left=Side(border_style='thin', color='CCCCCC'),
            right=Side(border_style='thin', color='CCCCCC'),
            top=Side(border_style='thin', color='CCCCCC'),
            bottom=Side(border_style='thin', color='CCCCCC')
        )
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Add headers
        headers = ['Criterion', 'Description', 'Points']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment

        def calculate_points(criterion):
            if not criterion.fulfilled:
                return 0.0
            elif criterion.custom_score is not None:
                return max(0.0, min(1.0, criterion.custom_score))
            else:
                return max(0.0, min(1.0, criterion.default_points))

        # Add criteria data
        current_row = 2
        for criterion in self.criteria:
            points = calculate_points(criterion)

            worksheet.cell(row=current_row, column=1, value=criterion.name)
            worksheet.cell(row=current_row, column=2, value=criterion.description)
            worksheet.cell(row=current_row, column=3, value=points)

            # Apply styling
            for col in range(1, 4):
                cell = worksheet.cell(row=current_row, column=col)
                cell.border = border
                cell.font = criterion_font
                cell.fill = criterion_fill
                cell.alignment = wrap_alignment
                if col == 3:  # Points column
                    cell.alignment = center_alignment

            current_row += 1

        # Set column widths
        column_widths = {
            'A': 25,  # Criterion
            'B': 45,  # Description
            'C': 10  # Points
        }

        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width

    def to_excel(self, filename: str) -> bytes:
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            self.to_excel_worksheet(writer, filename)

            # Remove default sheet if it exists
            if 'Sheet' in writer.book.sheetnames:
                writer.book.remove(writer.book['Sheet'])

        excel_buffer.seek(0)
        return excel_buffer.getvalue()
